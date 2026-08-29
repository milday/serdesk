import os
import sys
import time
import shutil
import openpyxl
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException, NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.support.ui import Select

# Keep downloaded drivers in the project so later runs work offline.
APP_DIR = os.path.dirname(os.path.abspath(__file__))
DRIVERS_DIR = os.path.join(APP_DIR, "drivers")


class ServiceDeskAutomation:
    def __init__(self):
        self.url = "https://servicedesk.adira.co.id/HEAT/"
        self.driver = None
        self.browser_type = "auto"
        self.location = ""
        self.setup_log = []
        self.role_already_selected = False
        self.detected_role = None

    # Any of these already shown as the dashboard role button means skip the picker.
    KNOWN_ROLES = (
        "L2 / L3 Analyst",
        "Mobile Self Service",
        "Self Service IT",
        "Self Service User",
    )
    ROLE_BUTTON_XPATH = (
        "//button[@type='button' and contains(@class, 'x-btn-text')]"
    )
    ROLE_PICKER_XPATH = "/html/body/div[4]/div/form/div[3]"
    ROLE_SUBMIT_XPATH = "/html/body/div[4]/div/form/div[4]/div/button"

    def _log_setup(self, message):
        print(message)
        self.setup_log.append(message)

    def _platform_key(self):
        if os.name == "nt":
            return "windows"
        if sys.platform == "darwin":
            return "macos"
        return "linux"

    def _driver_names(self, base_name):
        names = [base_name]
        if os.name == "nt":
            names.insert(0, base_name + ".exe")
        else:
            names.append(base_name + ".exe")
        return names

    def _ensure_driver_cache(self):
        os.makedirs(os.path.join(DRIVERS_DIR, self._platform_key()), exist_ok=True)
        selenium_cache = os.path.join(DRIVERS_DIR, "selenium-cache")
        os.makedirs(selenium_cache, exist_ok=True)
        # Selenium Manager (built into Selenium 4.6+) stores drivers here.
        os.environ["SE_CACHE_PATH"] = selenium_cache
        return selenium_cache

    def _find_existing_driver(self, base_name):
        names = self._driver_names(base_name)
        search_roots = [
            os.path.join(DRIVERS_DIR, self._platform_key()),
            DRIVERS_DIR,
        ]
        for root in search_roots:
            for name in names:
                path = os.path.join(root, name)
                if os.path.isfile(path):
                    return os.path.abspath(path)

        if os.path.isdir(DRIVERS_DIR):
            for dirpath, _, filenames in os.walk(DRIVERS_DIR):
                for name in names:
                    if name in filenames:
                        path = os.path.join(dirpath, name)
                        if os.path.isfile(path) and not path.lower().endswith(".zip"):
                            return os.path.abspath(path)

        for name in names:
            found = shutil.which(name)
            if found:
                return found
        return None

    def _find_browser_binary(self, candidates):
        for path in candidates:
            if path and os.path.isfile(path):
                return path
        return None

    def _find_firefox_binary(self):
        env_path = os.environ.get("FIREFOX_BINARY") or os.environ.get("FIREFOX_BIN")
        which = shutil.which("firefox") or shutil.which("firefox.exe")
        return self._find_browser_binary([
            env_path,
            which,
            r"C:\Program Files\Mozilla Firefox\firefox.exe",
            r"C:\Program Files (x86)\Mozilla Firefox\firefox.exe",
            os.path.expandvars(r"%LOCALAPPDATA%\Mozilla Firefox\firefox.exe"),
            "/usr/bin/firefox",
            "/usr/lib/firefox/firefox",
            "/snap/bin/firefox",
            "/Applications/Firefox.app/Contents/MacOS/firefox",
        ])

    def _find_chrome_binary(self):
        env_path = os.environ.get("CHROME_BINARY") or os.environ.get("CHROME_BIN")
        which = (
            shutil.which("google-chrome")
            or shutil.which("google-chrome-stable")
            or shutil.which("chrome")
            or shutil.which("chrome.exe")
            or shutil.which("chromium")
            or shutil.which("chromium-browser")
        )
        return self._find_browser_binary([
            env_path,
            which,
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
            "/usr/bin/google-chrome",
            "/usr/bin/google-chrome-stable",
            "/usr/bin/chromium",
            "/usr/bin/chromium-browser",
            "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
        ])

    def _cache_driver_locally(self, src_path, base_name):
        if not src_path or not os.path.isfile(src_path):
            return src_path
        dest_dir = os.path.join(DRIVERS_DIR, self._platform_key())
        os.makedirs(dest_dir, exist_ok=True)
        dest_name = self._driver_names(base_name)[0]
        dest_path = os.path.join(dest_dir, dest_name)
        try:
            if os.path.abspath(src_path) != os.path.abspath(dest_path):
                shutil.copy2(src_path, dest_path)
                if os.name != "nt":
                    os.chmod(dest_path, 0o755)
                self._log_setup(f"Cached driver in project folder: {dest_path}")
            return dest_path
        except Exception as e:
            self._log_setup(f"Could not copy driver into project folder: {e}")
            return src_path

    def _normalize_role_text(self, text):
        return " ".join((text or "").split())

    def get_active_dashboard_role(self):
        """Return the already-applied dashboard role name, or None.

        After login, HEAT may land on the dashboard with the current role already
        applied (ExtJS button, e.g. id=ext-gen50 class=x-btn-text). Match any known
        role from the picker list, not only L2 / L3 Analyst.
        """
        if not self.driver:
            return None

        known = {self._normalize_role_text(name): name for name in self.KNOWN_ROLES}

        try:
            for btn in self.driver.find_elements(By.XPATH, self.ROLE_BUTTON_XPATH):
                if not btn.is_displayed():
                    continue
                label = self._normalize_role_text(btn.text)
                if label in known:
                    return known[label]
        except Exception:
            pass

        # Fallback: recorded ExtJS id, only if the label is a known role.
        try:
            btn = self.driver.find_element(By.ID, "ext-gen50")
            if btn.is_displayed():
                label = self._normalize_role_text(btn.text)
                if label in known:
                    return known[label]
        except NoSuchElementException:
            pass

        return None

    def is_role_already_active(self):
        """True if dashboard already shows a known role button."""
        return self.get_active_dashboard_role() is not None

    def _role_picker_visible(self):
        """True if the post-login role picker form is on screen."""
        if not self.driver:
            return False
        try:
            el = self.driver.find_element(By.XPATH, self.ROLE_PICKER_XPATH)
            return el.is_displayed()
        except NoSuchElementException:
            return False

    def setup_driver(self, headless=False):
        """Initialize the WebDriver using local cache, installed browsers, then downloads."""
        self.setup_log = []
        self._ensure_driver_cache()
        requested = (self.browser_type or "auto").strip().lower()
        if requested == "auto":
            order = ["firefox", "chrome"]
        elif requested == "chrome":
            order = ["chrome", "firefox"]
        else:
            order = ["firefox", "chrome"]

        last_error = None
        for browser in order:
            self._log_setup(f"Setting up {browser.capitalize()} WebDriver...")
            try:
                ok = (
                    self._setup_firefox_driver(headless)
                    if browser == "firefox"
                    else self._setup_chrome_driver(headless)
                )
            except Exception as e:
                ok = False
                last_error = e
                self._log_setup(f"{browser.capitalize()} setup error: {e}")
            if ok:
                self.browser_type = browser
                return True
            if browser != order[-1]:
                self._log_setup(f"{browser.capitalize()} unavailable, trying the next browser...")

        if last_error:
            self._log_setup(f"Error in WebDriver setup: {last_error}")
        return False

    def _setup_firefox_driver(self, headless=False):
        """Initialize Firefox using bundled/local geckodriver, then cache/download if needed."""
        options = FirefoxOptions()
        if headless:
            options.add_argument("--headless")

        firefox_bin = self._find_firefox_binary()
        if firefox_bin:
            options.binary_location = firefox_bin
            self._log_setup(f"Using existing Firefox: {firefox_bin}")
        else:
            self._log_setup("Firefox browser not found on this computer.")
            return False

        errors = []

        local_driver = self._find_existing_driver("geckodriver")
        if local_driver:
            try:
                self._log_setup(f"Using local geckodriver: {local_driver}")
                self.driver = webdriver.Firefox(
                    service=FirefoxService(local_driver), options=options
                )
                self._log_setup("✅ Firefox WebDriver initialized (local driver)!")
                return True
            except Exception as e:
                errors.append(f"local geckodriver: {e}")
                self._log_setup(f"Local geckodriver failed: {e}")

        try:
            self._log_setup("Trying Selenium Manager (cache in project drivers/selenium-cache)...")
            self.driver = webdriver.Firefox(options=options)
            self._log_setup("✅ Firefox WebDriver initialized (Selenium Manager)!")
            return True
        except Exception as e:
            errors.append(f"Selenium Manager: {e}")
            self._log_setup(f"Selenium Manager Firefox failed: {e}")

        try:
            from webdriver_manager.core.driver_cache import DriverCacheManager
            from webdriver_manager.firefox import GeckoDriverManager

            cache_dir = os.path.join(DRIVERS_DIR, ".wdm")
            os.makedirs(cache_dir, exist_ok=True)
            # Reuse a previous project download without contacting GitHub.
            if self._find_existing_driver("geckodriver"):
                os.environ["WDM_LOCAL"] = "1"
            self._log_setup(f"Trying webdriver-manager cache at {cache_dir}...")
            gecko_path = GeckoDriverManager(
                cache_manager=DriverCacheManager(root_dir=DRIVERS_DIR)
            ).install()
            gecko_path = self._cache_driver_locally(gecko_path, "geckodriver")
            self.driver = webdriver.Firefox(
                service=FirefoxService(gecko_path), options=options
            )
            self._log_setup("✅ Firefox WebDriver initialized (downloaded to project folder)!")
            return True
        except Exception as e:
            errors.append(f"webdriver-manager: {e}")
            self._log_setup(f"Firefox setup failed: {e}")

        for err in errors:
            self._log_setup(f"  - {err}")
        return False

    def _setup_chrome_driver(self, headless=False):
        """Initialize Chrome using an installed browser and a project-local driver cache."""
        options = ChromeOptions()
        if headless:
            options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")

        chrome_bin = self._find_chrome_binary()
        if chrome_bin:
            options.binary_location = chrome_bin
            self._log_setup(f"Using existing Chrome: {chrome_bin}")
        else:
            self._log_setup("Chrome browser not found on this computer.")
            return False

        errors = []

        local_driver = self._find_existing_driver("chromedriver")
        if local_driver:
            try:
                self._log_setup(f"Using local chromedriver: {local_driver}")
                self.driver = webdriver.Chrome(
                    service=ChromeService(local_driver), options=options
                )
                self._log_setup("✅ Chrome WebDriver initialized (local driver)!")
                return True
            except Exception as e:
                errors.append(f"local chromedriver: {e}")
                self._log_setup(f"Local chromedriver failed: {e}")

        try:
            self._log_setup("Trying Selenium Manager (cache in project drivers/selenium-cache)...")
            self.driver = webdriver.Chrome(options=options)
            self._log_setup("✅ Chrome WebDriver initialized (Selenium Manager)!")
            return True
        except Exception as e:
            errors.append(f"Selenium Manager: {e}")
            self._log_setup(f"Selenium Manager Chrome failed: {e}")

        try:
            from webdriver_manager.chrome import ChromeDriverManager
            from webdriver_manager.core.driver_cache import DriverCacheManager

            os.makedirs(os.path.join(DRIVERS_DIR, ".wdm"), exist_ok=True)
            self._log_setup("Trying webdriver-manager for Chrome...")
            chrome_path = ChromeDriverManager(
                cache_manager=DriverCacheManager(root_dir=DRIVERS_DIR)
            ).install()
            chrome_path = self._cache_driver_locally(chrome_path, "chromedriver")
            self.driver = webdriver.Chrome(
                service=ChromeService(chrome_path), options=options
            )
            self._log_setup("✅ Chrome WebDriver initialized (downloaded to project folder)!")
            return True
        except Exception as e:
            errors.append(f"webdriver-manager: {e}")
            self._log_setup(f"Chrome setup failed: {e}")

        for err in errors:
            self._log_setup(f"  - {err}")
        return False

    def login(self, username, password):
        """Login to ServiceDesk"""
        try:
            print("Starting login process...")

            if not self.driver:
                print("❌ WebDriver not initialized")
                return False

            # Navigate to login page
            print("Navigating to login page...")
            self.driver.get(self.url)
            time.sleep(5)

            # Find and fill username
            try:
                username_field = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.ID, "UserName"))
                )
                username_field.clear()
                username_field.send_keys(username)
                print("✅ Username entered")
            except Exception as e:
                print(f"❌ Could not find username field: {str(e)}")
                return False

            # Find and fill password
            try:
                password_field = self.driver.find_element(By.ID, "Password")
                password_field.clear()
                password_field.send_keys(password)
                print("✅ Password entered")
            except Exception as e:
                print(f"❌ Could not find password field: {str(e)}")
                return False

            # Submit form
            try:
                submit_button = self.driver.find_element(By.XPATH, "//button[@type='submit']")
                submit_button.click()
                print("✅ Login form submitted")
            except Exception as e:
                print(f"❌ Could not submit form: {str(e)}")
                # Try alternative method
                password_field.send_keys(Keys.RETURN)
                print("✅ Login submitted via Enter key")

            # Wait for login to complete
            time.sleep(10)

            # Check if login was successful
            if "Access denied" in self.driver.page_source:
                print("❌ Login failed - Access denied")
                return False

            print("✅ Login completed!")
            return True

        except Exception as e:
            print(f"❌ Login error: {str(e)}")
            return False

    def select_role(self):
        """Select a role, or skip if any known role is already active on the dashboard."""
        try:
            if not self.driver:
                print("❌ WebDriver not initialized")
                return False

            self.role_already_selected = False
            self.detected_role = None
            print("Looking for role selection...")

            # Wait until either a dashboard role button or the role picker appears.
            try:
                WebDriverWait(self.driver, 15).until(
                    lambda d: self.is_role_already_active() or self._role_picker_visible()
                )
            except TimeoutException:
                print("⚠️ Neither an active role button nor role picker appeared yet")

            active_role = self.get_active_dashboard_role()
            if active_role:
                self.role_already_selected = True
                self.detected_role = active_role
                print(f"✅ Role '{active_role}' already detected — skipping role selection")
                return True

            # Try to find and click role
            try:
                role_div = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, self.ROLE_PICKER_XPATH))
                )
                role_div.click()
                print("✅ Role selected")
                time.sleep(2)
            except Exception as e:
                print(f"⚠️ Could not click role: {str(e)}")

            # Submit role selection
            try:
                submit_button = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, self.ROLE_SUBMIT_XPATH))
                )
                submit_button.click()
                print("✅ Role submitted")
            except Exception as e:
                print(f"⚠️ Could not submit role: {str(e)}")

            # Wait for dashboard
            print("Waiting for dashboard...")
            time.sleep(15)
            print("✅ Dashboard loaded!")
            return True

        except Exception as e:
            print(f"❌ Role selection error: {str(e)}")
            return False

    def navigate_to_new_incident(self):
        """Navigate to new incident form"""
        try:
            if not self.driver:
                print("❌ WebDriver not initialized")
                return False

            print("Looking for 'Report an Issue' button...")

            # Switch back to main content first
            try:
                self.driver.switch_to.default_content()
                print("✅ Switched to main content")
            except:
                pass

            # Wait for the page to load
            time.sleep(5)

            # Look for various possible selectors
            report_button_selectors = [
                "//button[contains(text(), 'Report an Issue')]",
                "//a[contains(text(), 'Report an Issue')]",
                "//button[contains(text(), 'Report Issue')]",
                "//a[contains(text(), 'Report Issue')]",
                "//button[@id='ext-gen60']",  # From the HTML source
                "//table[@id='ext-comp-1049']//button",
                "//button[contains(@class, 'x-btn-text') and contains(text(), 'Report')]"
            ]

            report_button = None
            for selector in report_button_selectors:
                try:
                    report_button = self.driver.find_element(By.XPATH, selector)
                    if report_button and report_button.is_displayed():
                        print(f"✅ Found Report button with selector: {selector}")
                        break
                except:
                    continue

            if report_button:
                report_button.click()
                print("✅ Clicked 'Report an Issue'")
            else:
                print("❌ Could not find 'Report an Issue' button")
                return False

            # Wait for form to load
            print("Waiting for incident form to load...")
            time.sleep(10)

            # Check for iframes and switch to the form
            try:
                iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
                print(f"Found {len(iframes)} iframes")

                if iframes:
                    # Switch to the largest iframe (likely the form)
                    largest_iframe = max(iframes, key=lambda x: x.size['width'] * x.size['height'])
                    self.driver.switch_to.frame(largest_iframe)
                    print("✅ Switched to incident form iframe")

                    # Wait longer for form elements to load completely
                    print("Waiting for form elements to load completely...")
                    time.sleep(10)

                    # Take a screenshot for debugging
                    try:
                        self.driver.save_screenshot("form_debug.png")
                        print("📸 Screenshot saved as 'form_debug.png'")
                    except:
                        pass

                    return True
                else:
                    print("❌ No iframes found")
                    return False

            except Exception as e:
                print(f"❌ Error switching to iframe: {str(e)}")
                return False

        except Exception as e:
            print(f"❌ Navigation error: {str(e)}")
            return False

    def _wait_for_field_validation(self, field_element, expected_value, max_wait=15):
        """Wait for dropdown field to be validated and populated"""
        try:
            print(f"    ⏳ Waiting for field validation (max {max_wait}s)...")
            start_time = time.time()

            while time.time() - start_time < max_wait:
                current_value = field_element.get_attribute("value") or ""

                # Check if the field has been populated with a valid value
                if expected_value.lower() in current_value.lower() and len(current_value) > 0:
                    print(f"    ✅ Field validated: '{current_value}'")
                    return True

                time.sleep(1)

            print(f"    ⚠️ Field validation timeout after {max_wait}s")
            return False

        except Exception as e:
            print(f"    ❌ Error waiting for field validation: {str(e)}")
            return False

    def fill_incident_form(self, subject, description):
        """Fill the incident form with proper timing and validation"""
        try:
            start_time = time.time()
            if not self.driver:
                print("❌ WebDriver not initialized")
                return False

            print("Filling incident form with proper timing...")

            # Wait for form to be ready
            print("Waiting for form to be ready...")
            time.sleep(3)

            print("\n=== USING TIMING-AWARE APPROACH ===")

            # Fill Service field with validation wait
            print("\n🎯 Filling Service field...")
            service_element = self.driver.find_element(By.XPATH, "//input[@id='ext-comp-1116']")
            service_element.clear()
            service_element.send_keys("Perangkat Kerja")
            print("    ✅ Typed 'Perangkat Kerja'")

            # Wait for validation
            if not self._wait_for_field_validation(service_element, "Perangkat Kerja"):
                print("❌ Service field validation failed")
                return False

            # Wait between fields
            time.sleep(3)

            # Fill System field with validation wait
            print("\n🎯 Filling System field...")
            system_element = self.driver.find_element(By.XPATH, "//input[@id='ext-comp-1110']")
            system_element.clear()
            system_element.send_keys("Software")
            print("    ✅ Typed 'Software'")

            # Wait for validation
            if not self._wait_for_field_validation(system_element, "Software"):
                print("❌ System field validation failed")
                return False

            # Fill Location field by tabbing once after System, when configured from GUI
            location_filled = False
            location_value = (self.location or "").strip()
            if location_value:
                print("\n🎯 Filling Location field...")
                system_element.send_keys(Keys.TAB)
                time.sleep(1)

                location_element = self.driver.switch_to.active_element
                location_element.clear()
                location_element.send_keys(location_value)
                print(f"    ✅ Typed Location '{location_value}'")
                location_filled = True
            else:
                print("\n🎯 Location field skipped (empty setting)")

            # Wait between fields
            time.sleep(3)

            # Fill Sub System field with validation wait
            print("\n🎯 Filling Sub System field...")
            subsystem_element = self.driver.find_element(By.XPATH, "//input[@id='ext-comp-1111']")
            subsystem_element.clear()
            subsystem_element.send_keys("Setting & Configuration")
            print("    ✅ Typed 'Setting & Configuration'")

            # Wait for validation
            if not self._wait_for_field_validation(subsystem_element, "Setting"):
                print("❌ Sub System field validation failed")
                return False

            # Wait before filling text fields
            time.sleep(3)

            # Fill Subject field
            print("\n🎯 Filling Subject field...")
            subject_element = self.driver.find_element(By.XPATH, "//input[@id='ext-comp-1106']")
            subject_element.clear()
            subject_element.send_keys(subject)
            print(f"✅ Subject field filled")

            # Wait before filling description
            time.sleep(2)

            # Fill Description field
            print("\n🎯 Filling Description field...")
            description_element = self.driver.find_element(By.XPATH, "//textarea[@id='ext-comp-1107']")
            description_element.clear()
            description_element.send_keys(description)
            print(f"✅ Description field filled")

            # Final wait before submission
            print("\n⏳ Waiting for all fields to stabilize before submission...")
            time.sleep(8)

            # Summary
            print(f"\n📋 FORM FILLING SUMMARY:")
            print(f"  Service Category: ⏭️ (Skipped - left as is)")
            print(f"  Service: ✅ (Perangkat Kerja)")
            print(f"  System: ✅ (Software)")
            print(f"  Location: {'✅ (' + location_value + ')' if location_filled else '⏭️ (Skipped - empty setting)'}")
            print(f"  Sub System: ✅ (Setting & Configuration)")
            print(f"  Current Location: ⏭️ (Skipped - left as is)")
            print(f"  Subject: ✅")
            print(f"  Description: ✅")

            elapsed_time = time.time() - start_time
            print(f"✅ Form filling completed! (took {elapsed_time:.1f} seconds)")
            return True

        except Exception as e:
            print(f"❌ Form filling error: {str(e)}")
            return False

    def submit_incident(self):
        """Submit the incident form and handle confirmation dialog"""
        try:
            if not self.driver:
                print("❌ WebDriver not initialized")
                return None

            print("Submitting incident form...")

            # Try to find submit button
            submit_selectors = [
                "//button[contains(text(), 'Submit')]",
                "//input[@value='Submit' or @value='Submit Incident']",
                "//input[@type='submit']",
                "//button[@type='submit']"
            ]

            submit_button = None
            for selector in submit_selectors:
                try:
                    submit_button = self.driver.find_element(By.XPATH, selector)
                    if submit_button and submit_button.is_displayed():
                        break
                except:
                    continue

            if submit_button:
                submit_button.click()
                print("✅ Submit button clicked")

                # Wait for confirmation dialog to appear
                time.sleep(3)

                # Handle confirmation dialog
                try:
                    print("Looking for confirmation dialog...")

                    # Try to find the "Yes" button in confirmation dialog
                    confirmation_selectors = [
                        "//button[contains(text(), 'Yes')]",
                        "//input[@value='Yes']",
                        "//button[@id and contains(text(), 'Yes')]",
                        "//div[contains(@class, 'x-window')]//button[contains(text(), 'Yes')]"
                    ]

                    yes_button = None
                    for selector in confirmation_selectors:
                        try:
                            yes_button = WebDriverWait(self.driver, 5).until(
                                EC.element_to_be_clickable((By.XPATH, selector))
                            )
                            if yes_button:
                                break
                        except:
                            continue

                    if yes_button:
                        yes_button.click()
                        print("✅ Confirmation 'Yes' button clicked")

                        # Wait for submission to complete
                        time.sleep(10)

                        # Try to get ticket number or confirmation from page
                        try:
                            page_text = self.driver.page_source
                            if 'incident' in page_text.lower() and ('submitted' in page_text.lower() or 'created' in page_text.lower()):
                                # Try to extract ticket number
                                import re
                                ticket_pattern = r'(?:incident|ticket)[\s#:]*([A-Z0-9-]+)'
                                match = re.search(ticket_pattern, page_text, re.IGNORECASE)
                                if match:
                                    return f"SUBMITTED_SUCCESS_{match.group(1)}"
                                else:
                                    return "SUBMITTED_SUCCESS"
                            else:
                                return "SUBMITTED_UNKNOWN"
                        except:
                            return "SUBMITTED_SUCCESS"
                    else:
                        print("❌ Could not find confirmation 'Yes' button")
                        return None

                except Exception as e:
                    print(f"⚠️ Error handling confirmation dialog: {str(e)}")
                    # Maybe the dialog didn't appear, submission might have worked
                    time.sleep(5)
                    return "SUBMITTED_UNKNOWN"

            else:
                print("❌ Could not find submit button")
                return None

        except Exception as e:
            print(f"❌ Submit error: {str(e)}")
            return None

    def navigate_back_to_dashboard(self):
        """Navigate back to dashboard"""
        try:
            if not self.driver:
                print("❌ WebDriver not initialized")
                return False

            print("Navigating back to dashboard...")

            # Switch back to main content
            self.driver.switch_to.default_content()

            # Go to dashboard URL
            dashboard_url = "https://servicedesk.adira.co.id/HEAT/Default.aspx"
            self.driver.get(dashboard_url)
            time.sleep(5)

            print("✅ Returned to dashboard")
            return True

        except Exception as e:
            print(f"❌ Navigation error: {str(e)}")
            return False

    def process_tickets(self):
        """Process tickets from Excel file"""
        try:
            print("Loading Excel file: tickets.xlsx")

            if not os.path.exists('tickets.xlsx'):
                print("❌ tickets.xlsx file not found!")
                return False

            # Load workbook for reading values
            print("Loading Excel file for reading values...")
            workbook_read = openpyxl.load_workbook('tickets.xlsx', data_only=True)
            worksheet_read = workbook_read.active

            # Load workbook for saving (preserve formulas)
            print("Loading Excel file for preserving formulas...")
            workbook_save = openpyxl.load_workbook('tickets.xlsx', data_only=False)
            worksheet_save = workbook_save.active

            if not worksheet_read or not worksheet_save:
                print("❌ Could not access worksheets")
                return False

            print(f"✅ Excel files loaded")
            print(f"Total rows: {worksheet_read.max_row}")

            successful_tickets = 0
            failed_tickets = 0
            results = []

            # Process each row (skip header)
            for row_num in range(2, worksheet_read.max_row + 1):
                try:
                    # Get data from Excel (using the read workbook for calculated values)
                    subject_cell = worksheet_read.cell(row=row_num, column=1)
                    description_cell = worksheet_read.cell(row=row_num, column=2)

                    # Debug: Show raw cell values
                    print(f"🔍 Row {row_num} - Raw values:")
                    print(f"  Subject cell: '{subject_cell.value}' (type: {type(subject_cell.value)})")
                    print(f"  Description cell: '{description_cell.value}' (type: {type(description_cell.value)})")

                    subject = str(subject_cell.value or "").strip()
                    description = str(description_cell.value or "").strip()

                    print(f"  Processed - Subject: '{subject}', Description: '{description}'")

                    if not subject or not description:
                        print(f"⚠️ Skipping row {row_num}: Empty data")
                        continue

                    print(f"\n📋 Processing ticket {row_num - 1}")
                    print(f"Subject: {subject}")
                    print(f"Description: {description}")

                    # Navigate to new incident form
                    if not self.navigate_to_new_incident():
                        print("❌ Failed to navigate to form")
                        failed_tickets += 1
                        results.append(f"Ticket {row_num - 1}: FAILED - Navigation failed")
                        continue

                    # Fill form
                    if not self.fill_incident_form(subject, description):
                        print("❌ Failed to fill form")
                        failed_tickets += 1
                        results.append(f"Ticket {row_num - 1}: FAILED - Form filling failed")
                        continue

                    # Submit form
                    ticket_number = self.submit_incident()

                    if ticket_number:
                        print(f"✅ Success! Ticket: {ticket_number}")
                        successful_tickets += 1
                        results.append(f"Ticket {row_num - 1}: SUCCESS - {ticket_number}")

                        # Update Excel with ticket number (using the save workbook to preserve formulas)
                        worksheet_save.cell(row=row_num, column=3, value=ticket_number)

                        # Navigate back for next ticket
                        if row_num < worksheet_read.max_row:
                            self.navigate_back_to_dashboard()
                            time.sleep(3)
                    else:
                        print("❌ Submission failed")
                        failed_tickets += 1
                        results.append(f"Ticket {row_num - 1}: FAILED - Submission failed")

                    # Save Excel after each ticket (using the save workbook to preserve formulas)
                    workbook_save.save('tickets.xlsx')

                except Exception as e:
                    print(f"❌ Error processing ticket {row_num - 1}: {str(e)}")
                    failed_tickets += 1
                    results.append(f"Ticket {row_num - 1}: FAILED - {str(e)}")

            # Final save (using the save workbook to preserve formulas)
            workbook_save.save('tickets.xlsx')

            # Close workbooks to release file handles
            workbook_read.close()
            workbook_save.close()

            # Write results
            with open('ticket_results.txt', 'w', encoding='utf-8') as f:
                f.write("SERVICEDESK TICKET PROCESSING RESULTS\n")
                f.write("=" * 50 + "\n")
                f.write(f"Total processed: {successful_tickets + failed_tickets}\n")
                f.write(f"Successful: {successful_tickets}\n")
                f.write(f"Failed: {failed_tickets}\n")
                if successful_tickets + failed_tickets > 0:
                    success_rate = (successful_tickets / (successful_tickets + failed_tickets)) * 100
                    f.write(f"Success rate: {success_rate:.1f}%\n\n")

                f.write("DETAILED RESULTS:\n")
                for result in results:
                    f.write(result + "\n")

            print(f"\n🎯 FINAL SUMMARY")
            print(f"Total tickets: {successful_tickets + failed_tickets}")
            print(f"Successful: {successful_tickets}")
            print(f"Failed: {failed_tickets}")
            if successful_tickets + failed_tickets > 0:
                success_rate = (successful_tickets / (successful_tickets + failed_tickets)) * 100
                print(f"Success rate: {success_rate:.1f}%")

            return successful_tickets > 0

        except Exception as e:
            print(f"❌ Processing error: {str(e)}")
            return False
        finally:
            # Ensure workbooks are closed even if there's an error
            try:
                if 'workbook_read' in locals():
                    workbook_read.close()
                if 'workbook_save' in locals():
                    workbook_save.close()
            except:
                pass

    def run(self):
        """Run the complete automation"""
        try:
            print("=== ServiceDesk Automation Started ===")

            # Setup WebDriver
            if not self.setup_driver():
                print("❌ Failed to setup WebDriver")
                input("❌ Press Enter to continue...")  # Pause for error
                return False

            # Get credentials
            username = input("Enter ServiceDesk username: ")
            password = input("Enter ServiceDesk password: ")

            # Login
            if not self.login(username, password):
                print("❌ Login failed")
                input("❌ Press Enter to continue...")  # Pause for error
                return False

            # Select role
            if not self.select_role():
                print("❌ Role selection failed")
                input("❌ Press Enter to continue...")  # Pause for error
                return False

            # Process tickets
            if not self.process_tickets():
                print("❌ Ticket processing failed")
                input("❌ Press Enter to continue...")  # Pause for error
                return False

            print("✅ Automation completed successfully!")
            input("✅ Press Enter to close...")  # Pause for success
            return True

        except Exception as e:
            print(f"❌ Automation error: {str(e)}")
            print("❌ Full error details:")
            import traceback
            traceback.print_exc()
            input("❌ Press Enter to close after error...")  # Pause for error
            return False
        finally:
            # Close browser
            if self.driver:
                try:
                    self.driver.quit()
                    print("✅ Browser closed")
                except:
                    pass

if __name__ == "__main__":
    automation = ServiceDeskAutomation()
    automation.run()


